import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

def preparar_datos(datos):
    # Asegúrate de que todas las entradas en las columnas que podrían tener tipos mixtos sean cadenas.
    columnas_a_string = ['DESCRIPCIÓN', 'MARCA', 'MODELO', 'P/N', 'S/N', 'OBSERVACIONES', 'STATUS', 'UBICACIÓN', 'MEDIDA']
    for columna in columnas_a_string:
        datos[columna] = datos[columna].astype(str)
    
    # Reemplaza las cadenas vacías generadas por la conversión a `str` con un valor que indique que no hay datos
    datos.replace({'': 'No disponible'}, inplace=True)

    # Asegúrate de que las columnas numéricas, como 'PRECIO UNIT' y 'TOTAL', no tengan valores no numéricos.
    columnas_numericas = ['CANT.', 'PRECIO UNIT', 'TOTAL']
    for columna in columnas_numericas:
        datos[columna] = pd.to_numeric(datos[columna], errors='coerce').fillna(0)
    
    return datos

def cargar_datos(archivo):
    if archivo is not None:
        if archivo.name.endswith('.xlsx'):
            # Captura el nombre de la primera hoja para poder manipularla más adelante
            excel_file = pd.ExcelFile(archivo)
            sheet_name = excel_file.sheet_names[0]
            st.session_state['hoja'] = sheet_name
            datos = pd.read_excel(excel_file, sheet_name=sheet_name, skiprows=3)  # Ajuste para encabezados en la fila 4
        elif archivo.name.endswith('.csv'):
            st.session_state['hoja'] = None
            datos = pd.read_csv(archivo)
        datos = preparar_datos(datos)
        # Omitir columnas completamente vacías
        datos.dropna(axis=1, how='all', inplace=True)
        return datos
    return None

# Convierte un DataFrame en bytes para descargarlo como archivo Excel
def convertir_a_excel(df, original_bytes=None, filas_originales=0, sheet_name=None):
    """Convierte un DataFrame en bytes tomando como base el archivo original.

    Si se proporciona ``original_bytes`` se preserva el documento original y
    solo se añaden las filas nuevas, manteniendo el formato y añadiendo
    bordes a las nuevas celdas.

    ``sheet_name`` permite indicar la hoja sobre la que se añadirán los datos
    cuando se parte de un documento existente.
    """

    if original_bytes is not None:
        wb = load_workbook(filename=BytesIO(original_bytes))
        # Si se especifica un nombre de hoja, se selecciona esa hoja; de lo contrario, se utiliza la activa
        ws = wb[sheet_name] if sheet_name else wb.active
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in df.iloc[filas_originales:].fillna("").itertuples(index=False):
            ws.append(list(row))
            for col_idx, _ in enumerate(row, start=1):
                cell = ws.cell(row=ws.max_row, column=col_idx)
                cell.border = border

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.fillna("").to_excel(writer, index=False)
    buffer.seek(0)
    return buffer

# Función para resaltar coincidencias en los datos (simplificada para Streamlit)
def resaltar_coincidencias(datos, texto_busqueda):
    if datos is not None and texto_busqueda != "":
        return datos[datos.apply(lambda row: texto_busqueda.lower() in row.to_string().lower(), axis=1)]
    return pd.DataFrame()

# Función para generar el resumen de la búsqueda
def resumen_busqueda(datos_filtrados):
    if datos_filtrados.empty:
        return "No se encontraron resultados."
    
    try:
        # Encuentra la cantidad máxima y mínima y sus respectivos ítems
        max_cant = datos_filtrados['CANT.'].max()
        min_cant = datos_filtrados['CANT.'].min()
        max_item_row = datos_filtrados[datos_filtrados['CANT.'] == max_cant].iloc[0]
        min_item_row = datos_filtrados[datos_filtrados['CANT.'] == min_cant].iloc[0]

        # Extrae los detalles del ítem con la cantidad máxima
        max_item = max_item_row['ITEM']
        max_precio = max_item_row['PRECIO UNIT']
        ubicacion_max = max_item_row['UBICACIÓN']
        status_max = max_item_row['STATUS']

        # Extrae los detalles del ítem con la cantidad mínima
        min_item = min_item_row['ITEM']
        min_precio = min_item_row['PRECIO UNIT']
        ubicacion_min = min_item_row['UBICACIÓN']
        status_min = min_item_row['STATUS']

        # Calcula el precio promedio y el rango de precios
        precio_promedio = datos_filtrados['PRECIO UNIT'].mean()
        rango_precios = (datos_filtrados['PRECIO UNIT'].min(), datos_filtrados['PRECIO UNIT'].max())

        # Crea el string de resumen
        resumen_str = f"""
El item con más productos es {max_item} con una cantidad de {max_cant} y un precio de {max_precio}. Ubicación: {ubicacion_max}, Status: {status_max}
El item con menos productos es {min_item} con una cantidad de {min_cant} y un precio de {min_precio}. Ubicación: {ubicacion_min}, Status: {status_min}
Total de productos encontrados: {len(datos_filtrados)}
Precio promedio: {precio_promedio:.2f}
Rango de precios: {rango_precios[0]} - {rango_precios[1]}
        """
    except Exception as e:
        resumen_str = f"Error al generar resumen: {e}"
    
    return resumen_str

# Streamlit App
def app():
    st.title("Visualizador, buscador y generador de resumen de inventario")

    # Descripción de la aplicación con st.write
    st.write("""
    Esta aplicación permite cargar archivos de inventario en formatos CSV o XLSX, 
    visualizar los datos y buscar dentro del inventario. También puedes generar un 
    resumen del inventario que incluye el item con mayor y menor cantidad, el precio 
    promedio y el rango de precios.
    
    **Instrucciones:**
    - Utiliza el botón 'Cargar archivo' para subir tu archivo de inventario.
    - Escribe en el cuadro de búsqueda para filtrar los resultados.
    - El resumen del inventario se generará automáticamente basado en los datos filtrados.
    """)
    

    if 'datos' not in st.session_state:
        st.session_state['datos'] = None
        st.session_state['archivo_bytes'] = None
        st.session_state['extension'] = None
        st.session_state['filas_originales'] = 0
        st.session_state['hoja'] = None

    archivo = st.file_uploader("Cargar archivo XLSX/CSV", type=['xlsx', 'csv'])
    if archivo is not None:
        st.session_state['datos'] = cargar_datos(archivo)
        st.session_state['archivo_bytes'] = archivo.getvalue()
        st.session_state['extension'] = '.xlsx' if archivo.name.endswith('.xlsx') else '.csv'
        st.session_state['filas_originales'] = len(st.session_state['datos'])

    datos = st.session_state['datos']

    if datos is not None:
        st.write("Datos Cargados:")
        st.dataframe(datos)

        with st.expander("Añadir nuevo item"):
            with st.form("form-nuevo-item"):
                col1, col2 = st.columns(2)
                item = col1.text_input('ITEM')
                descripcion = col1.text_input('DESCRIPCIÓN')
                marca = col1.text_input('MARCA')
                modelo = col1.text_input('MODELO')
                pn = col1.text_input('P/N')
                sn = col1.text_input('S/N')
                observ = col2.text_input('OBSERVACIONES')
                status = col2.text_input('STATUS')
                ubicacion = col2.text_input('UBICACIÓN')
                medida = col2.text_input('MEDIDA')
                cant = col1.number_input('CANT.', value=0, step=1)
                precio = col1.number_input('PRECIO UNIT', value=0.0)
                total = col1.number_input('TOTAL', value=0.0)
                enviado = st.form_submit_button('Añadir')

            if enviado:
                nuevo = {
                    'ITEM': item,
                    'DESCRIPCIÓN': descripcion,
                    'MARCA': marca,
                    'MODELO': modelo,
                    'P/N': pn,
                    'S/N': sn,
                    'OBSERVACIONES': observ,
                    'STATUS': status,
                    'UBICACIÓN': ubicacion,
                    'MEDIDA': medida,
                    'CANT.': cant,
                    'PRECIO UNIT': precio,
                    'TOTAL': total,
                }
                st.session_state['datos'] = pd.concat([
                    datos,
                    pd.DataFrame([nuevo])
                ], ignore_index=True)
                st.session_state['datos'].fillna('', inplace=True)
                datos = st.session_state['datos']
                st.success('Item añadido correctamente')

        texto_busqueda = st.text_input("Buscar texto en los datos:")
        if texto_busqueda:
            datos_filtrados = resaltar_coincidencias(datos, texto_busqueda)
            if not datos_filtrados.empty:
                st.write("Resultados de la búsqueda:")
                st.dataframe(datos_filtrados)
                st.write("Resumen de la búsqueda:")
                st.info(resumen_busqueda(datos_filtrados))
            else:
                st.warning("No se encontraron coincidencias.")

        if st.session_state.get('extension') == '.xlsx':
            excel_bytes = convertir_a_excel(
                datos,
                st.session_state.get('archivo_bytes'),
                st.session_state.get('filas_originales', 0),
                sheet_name=st.session_state.get('hoja'),
            )
            file_name = "inventario_actualizado.xlsx"
        else:
            tmp = BytesIO()
            datos.fillna('').to_csv(tmp, index=False)
            tmp.seek(0)
            excel_bytes = tmp
            file_name = "inventario_actualizado.csv"

        with st.expander("Previsualizar archivo"):
            if st.session_state.get('extension') == '.xlsx':
                preview_df = pd.read_excel(
                    BytesIO(excel_bytes.getvalue()),
                    sheet_name=st.session_state.get('hoja')
                )
            else:
                preview_df = pd.read_csv(BytesIO(excel_bytes.getvalue()))
            st.dataframe(preview_df)

        st.download_button(
            "Descargar inventario actualizado",
            data=excel_bytes,
            file_name=file_name,
        )

if __name__ == "__main__":
    app()

